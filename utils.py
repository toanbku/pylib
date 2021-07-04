import io
from openpyxl import load_workbook

class File_Interact():
  def __init__(self, file_path):
    self.file_path = file_path

  def read_file(self):
    f = io.open(self.file_path, 'r', encoding='utf-8')
    return f.read()

  def read_file_list(self):
    f = io.open(self.file_path, 'r', encoding='utf-8')
    l = f.read().split('\n')
    return l

  def write_file(self, content):
    f = io.open(self.file_path, 'w', encoding='utf-8')
    f.write(content)
    f.close()

  def write_file_list(self, content):
    f = io.open(self.file_path, 'w', encoding='utf-8')
    f.write('\n'.join(content))
    f.close()

  def read_cell(self, sheet_name, cell_name):
    wb = load_workbook(self.file_path)
    sheet_ranges = wb[sheet_name]
    return sheet_ranges[cell_name].value

  def update_cell(self, sheet_name, cell_name, content):
    wb = load_workbook(self.file_path)
    sheet_ranges = wb[sheet_name]
    sheet_ranges[cell_name].value = content
    wb.close()
    wb.save(self.file_path)


if __name__ == '__main__':
  file_interact = File_Interact(
      file_path='res.xlsx'
  )
  file1 = File_Interact(
    file_path='file1.txt'
  )
  file2 = File_Interact(
    file_path='file2.txt'
  )
  sheet_name = 'Sheet1'
  cell_name = 'A3'
  content = 'hihi'
  file_interact.update_cell(sheet_name, 'A1', 'Tên')
  file_interact.update_cell(sheet_name, 'B1', 'Tuổi')
  L1 = file1.read_file_list()
  L2 = file2.read_file_list()

  for i in range(0, len(L1)):
    cell_name = 'A%s'%(i+2)
    cell_age = 'B%s'%(i+2)
    file_interact.update_cell(sheet_name, cell_name, L1[i])
    file_interact.update_cell(sheet_name, cell_age, L2[i])
  print('done')
